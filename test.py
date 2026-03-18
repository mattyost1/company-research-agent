import anthropic
import json
from tavily import TavilyClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

client = anthropic.Anthropic(api_key="your-anthropic-key-here")
tavily = TavilyClient(api_key="your-tavily-key-here")

# --- TOOLS ---

tools = [
    {
        "name": "search_web",
        "description": "Search the web for financial information about a company",
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "The search query"}
            },
            "required": ["query"]
        }
    },
    {
        "name": "build_dcf_excel",
        "description": "Build a formatted DCF model in Excel once all financial data has been gathered",
        "input_schema": {
            "type": "object",
            "properties": {
                "company_name": {"type": "string"},
                "revenue": {"type": "number", "description": "Most recent annual revenue in $M"},
                "revenue_growth": {"type": "number", "description": "Revenue growth rate as a percentage e.g. 25 for 25%"},
                "ebitda_margin": {"type": "number", "description": "EBITDA margin as a percentage"},
                "tax_rate": {"type": "number", "description": "Effective tax rate as a percentage"},
                "capex_pct": {"type": "number", "description": "Capex as percentage of revenue"},
                "wacc": {"type": "number", "description": "WACC as a percentage"},
                "terminal_growth": {"type": "number", "description": "Terminal growth rate as a percentage"},
                "share_price": {"type": "number", "description": "Current share price in $"},
                "shares_outstanding": {"type": "number", "description": "Shares outstanding in millions"},
                "total_debt": {"type": "number", "description": "Total debt in $M"},
                "cash": {"type": "number", "description": "Cash and equivalents in $M"}
            },
            "required": [
                "company_name", "revenue", "revenue_growth", "ebitda_margin",
                "tax_rate", "capex_pct", "wacc", "terminal_growth",
                "share_price", "shares_outstanding", "total_debt", "cash"
            ]
        }
    }
]

# --- TOOL FUNCTIONS ---

def search_web(query):
    print(f"  Searching: {query}...")
    results = tavily.search(query=query)
    return results['results'][0]['content']

def build_dcf_excel(data):
    print(f"\n  Building Excel model for {data['company_name']}...")

    wb = Workbook()
    ws = wb.active
    ws.title = "DCF Model"

    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E79")
    blue_font = Font(name="Arial", color="0000FF", size=10)
    black_font = Font(name="Arial", color="000000", size=10)
    bold_black = Font(name="Arial", color="000000", bold=True, size=10)
    section_fill = PatternFill("solid", start_color="D6E4F0")
    highlight_fill = PatternFill("solid", start_color="FFF2CC")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(row, col, value, font=None, fill=None, align="left", num_format=None):
        c = ws.cell(row=row, column=col, value=value)
        if font: c.font = font
        if fill: c.fill = fill
        if num_format: c.number_format = num_format
        c.alignment = Alignment(horizontal=align)
        c.border = border
        return c

    # Column widths
    ws.column_dimensions["A"].width = 32
    for col in ["B","C","D","E","F","G"]:
        ws.column_dimensions[col].width = 14

    years = [2025, 2026, 2027, 2028, 2029]

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"].value = f"{data['company_name']} — DCF Valuation Model"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # --- ASSUMPTIONS ---
    row = 3
    ws.merge_cells(f"A{row}:G{row}")
    cell(row, 1, "KEY ASSUMPTIONS", header_font, header_fill, "center")

    assumptions = [
        ("Revenue Growth Rate", data["revenue_growth"] / 100, "0.0%"),
        ("EBITDA Margin", data["ebitda_margin"] / 100, "0.0%"),
        ("Tax Rate", data["tax_rate"] / 100, "0.0%"),
        ("Capex % of Revenue", data["capex_pct"] / 100, "0.0%"),
        ("WACC", data["wacc"] / 100, "0.0%"),
        ("Terminal Growth Rate", data["terminal_growth"] / 100, "0.0%"),
    ]
    assumption_rows = {}
    for label, value, fmt in assumptions:
        row += 1
        cell(row, 1, label, black_font)
        cell(row, 2, value, blue_font, num_format=fmt)
        assumption_rows[label] = row

    wacc_cell = f"$B${assumption_rows['WACC']}"
    tgr_cell = f"$B${assumption_rows['Terminal Growth Rate']}"
    growth_cell = f"$B${assumption_rows['Revenue Growth Rate']}"
    margin_cell = f"$B${assumption_rows['EBITDA Margin']}"
    tax_cell = f"$B${assumption_rows['Tax Rate']}"
    capex_cell = f"$B${assumption_rows['Capex % of Revenue']}"

    # --- PROJECTIONS ---
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    cell(row, 1, "FINANCIAL PROJECTIONS ($M)", header_font, header_fill, "center")

    row += 1
    cell(row, 1, "Metric", bold_black, section_fill)
    for i, yr in enumerate(years):
        cell(row, i + 2, str(yr), bold_black, section_fill, "center")

    revenue_row = row + 1
    ebitda_row = row + 2
    nopat_row = row + 3
    capex_row = row + 4
    fcf_row = row + 5
    pv_row = row + 6

    # Revenue
    row += 1
    cell(row, 1, "Revenue ($M)", black_font)
    base = data["revenue"]
    for i in range(5):
        val = round(base * ((1 + data["revenue_growth"] / 100) ** (i + 1)), 1)
        c = ws.cell(row=row, column=i + 2, value=val)
        c.font = blue_font if i == 0 else black_font
        c.number_format = '#,##0.0'
        c.border = border

    # EBITDA
    row += 1
    cell(row, 1, "EBITDA ($M)", black_font)
    for i in range(5):
        ref = ws.cell(row=revenue_row, column=i + 2).coordinate
        c = ws.cell(row=row, column=i + 2)
        c.value = f"={ref}*{margin_cell}"
        c.font = black_font
        c.number_format = '#,##0.0'
        c.border = border

    # NOPAT
    row += 1
    cell(row, 1, "NOPAT ($M)", black_font)
    for i in range(5):
        ref = ws.cell(row=ebitda_row, column=i + 2).coordinate
        c = ws.cell(row=row, column=i + 2)
        c.value = f"={ref}*(1-{tax_cell})"
        c.font = black_font
        c.number_format = '#,##0.0'
        c.border = border

    # Capex
    row += 1
    cell(row, 1, "Capex ($M)", black_font)
    for i in range(5):
        ref = ws.cell(row=revenue_row, column=i + 2).coordinate
        c = ws.cell(row=row, column=i + 2)
        c.value = f"=-{ref}*{capex_cell}"
        c.font = black_font
        c.number_format = '#,##0.0'
        c.border = border

    # FCF
    row += 1
    cell(row, 1, "Free Cash Flow ($M)", bold_black)
    for i in range(5):
        n = ws.cell(row=nopat_row, column=i + 2).coordinate
        k = ws.cell(row=capex_row, column=i + 2).coordinate
        c = ws.cell(row=row, column=i + 2)
        c.value = f"={n}+{k}"
        c.font = bold_black
        c.number_format = '#,##0.0'
        c.border = border

    # PV of FCFs
    row += 1
    cell(row, 1, "PV of FCF ($M)", black_font)
    for i in range(5):
        fcf = ws.cell(row=fcf_row, column=i + 2).coordinate
        c = ws.cell(row=row, column=i + 2)
        c.value = f"={fcf}/(1+{wacc_cell})^{i+1}"
        c.font = black_font
        c.number_format = '#,##0.0'
        c.border = border

    # --- VALUATION ---
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    cell(row, 1, "VALUATION SUMMARY", header_font, header_fill, "center")

    pv_start = ws.cell(row=pv_row, column=2).coordinate
    pv_end = ws.cell(row=pv_row, column=6).coordinate
    last_fcf = ws.cell(row=fcf_row, column=6).coordinate

    row += 1
    cell(row, 1, "Sum of PV FCFs ($M)", black_font)
    c = ws.cell(row=row, column=2)
    c.value = f"=SUM({pv_start}:{pv_end})"
    c.font = black_font
    c.number_format = '#,##0.0'
    c.border = border
    sum_pv = c.coordinate

    row += 1
    cell(row, 1, "Terminal Value ($M)", black_font)
    c = ws.cell(row=row, column=2)
    c.value = f"=({last_fcf}*(1+{tgr_cell}))/({wacc_cell}-{tgr_cell})"
    c.font = black_font
    c.number_format = '#,##0.0'
    c.border = border
    tv = c.coordinate

    row += 1
    cell(row, 1, "PV of Terminal Value ($M)", black_font)
    c = ws.cell(row=row, column=2)
    c.value = f"={tv}/(1+{wacc_cell})^5"
    c.font = black_font
    c.number_format = '#,##0.0'
    c.border = border
    pv_tv = c.coordinate

    row += 1
    cell(row, 1, "Enterprise Value ($M)", bold_black)
    c = ws.cell(row=row, column=2)
    c.value = f"={sum_pv}+{pv_tv}"
    c.font = bold_black
    c.number_format = '#,##0.0'
    c.border = border
    ev = c.coordinate

    row += 1
    cell(row, 1, "Less: Total Debt ($M)", black_font)
    cell(row, 2, data["total_debt"], blue_font, num_format='#,##0.0')
    debt = ws.cell(row=row, column=2).coordinate

    row += 1
    cell(row, 1, "Plus: Cash ($M)", black_font)
    cell(row, 2, data["cash"], blue_font, num_format='#,##0.0')
    cash = ws.cell(row=row, column=2).coordinate

    row += 1
    cell(row, 1, "Equity Value ($M)", bold_black)
    c = ws.cell(row=row, column=2)
    c.value = f"={ev}-{debt}+{cash}"
    c.font = bold_black
    c.number_format = '#,##0.0'
    c.border = border
    equity = c.coordinate

    row += 1
    cell(row, 1, "Shares Outstanding (M)", black_font)
    cell(row, 2, data["shares_outstanding"], blue_font, num_format='#,##0.0')
    shares = ws.cell(row=row, column=2).coordinate

    row += 1
    cell(row, 1, "Implied Share Price", bold_black, highlight_fill)
    c = ws.cell(row=row, column=2)
    c.value = f"={equity}/{shares}"
    c.font = Font(name="Arial", bold=True, size=12, color="000000")
    c.number_format = '$#,##0.00'
    c.fill = highlight_fill
    c.border = border
    implied = c.coordinate

    row += 1
    cell(row, 1, "Current Share Price", black_font)
    cell(row, 2, data["share_price"], blue_font, num_format='$#,##0.00')
    current = ws.cell(row=row, column=2).coordinate

    row += 1
    cell(row, 1, "Upside / Downside", bold_black)
    c = ws.cell(row=row, column=2)
    c.value = f"=({implied}-{current})/{current}"
    c.font = bold_black
    c.number_format = '0.0%'
    c.border = border

    filename = f"{data['company_name'].replace(' ', '_')}_DCF.xlsx"
    wb.save(filename)
    return f"DCF model saved as {filename}"

# --- AGENT LOOP ---

def run_agent(user_message):
    print(f"\nTask: {user_message}\n")
    messages = [{"role": "user", "content": user_message}]

    system = """You are a financial analyst agent. When asked to build a DCF model:
1. Search the web multiple times to gather all required financial data
2. Once you have enough data, call build_dcf_excel with the inputs
3. Use your best judgment on inputs where data is unclear
Do not ask for clarification — just research and build."""

    while True:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=4096,
            system=system,
            tools=tools,
            messages=messages
        )

        if response.stop_reason == "end_turn":
            print("\n" + response.content[0].text)
            break

        if response.stop_reason == "tool_use":
            messages.append({"role": "assistant", "content": response.content})
            tool_results = []

            for block in response.content:
                if block.type == "tool_use":
                    if block.name == "search_web":
                        result = search_web(block.input["query"])
                    elif block.name == "build_dcf_excel":
                        result = build_dcf_excel(block.input)

                    tool_results.append({
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": result
                    })

            messages.append({"role": "user", "content": tool_results})

# --- RUN ---
run_agent("Build me a DCF model for Palantir and save it as an Excel file.")